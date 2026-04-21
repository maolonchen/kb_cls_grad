import pandas as pd
import openpyxl
from openpyxl.utils import range_boundaries
from collections import defaultdict
import tempfile
import os

def fill_merged_cells(file_path, output_path, sheet_names=None):
    """
    填充Excel工作簿中指定工作表的合并单元格
    
    参数:
        file_path: Excel文件路径
        sheet_names: 要处理的工作表名称列表，如果为None则处理所有工作表
    """
    # 加载工作簿
    wb = openpyxl.load_workbook(file_path)
    
    # 如果未指定工作表，则处理所有工作表
    if sheet_names is None:
        sheet_names = wb.sheetnames
    
    # 处理每个工作表
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        print(f"正在处理工作表: {sheet_name}")
        
        # 获取所有合并单元格范围（创建副本）
        merged_ranges = list(ws.merged_cells.ranges)
        
        # 遍历所有合并单元格
        for merged_range in merged_ranges:
            # 获取合并单元格的边界
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            
            # 获取主单元格的值(左上角单元格)
            main_cell = ws.cell(row=min_row, column=min_col)
            value = main_cell.value
            
            # 先取消合并
            ws.unmerge_cells(str(merged_range))
            
            # 填充合并区域中的所有单元格
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = value
    
    # 保存修改后的文件
    wb.save(output_path)
    print(f"处理完成，结果已保存为: {output_path}")
    return output_path

def number_to_letters(n):
    """
    将数字转换为字母序列
    """
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

def process_hierarchical_table_generic(df, target_col='对应特征'):
    """
    通用层级表格编号处理函数
    
    参数:
    df: pandas DataFrame, 包含表格数据
    target_col: str, 目标列名称，此列之前的列都需要编号
    
    返回:
    处理后的DataFrame
    """
    # 创建结果DataFrame
    result_df = df.copy()
    
    # 确定需要编号的列
    if target_col in df.columns:
        target_idx = df.columns.get_loc(target_col)
        num_cols = target_idx
    else:
        num_cols = len(df.columns) - 1
    
    # 存储每一行的父级编号
    parent_codes = [""] * len(df)
    
    # 存储每个父级路径下的子项计数器
    counters = defaultdict(lambda: defaultdict(int))
    
    # 处理每一列
    for col_idx in range(num_cols):
        # 获取当前列数据
        col_data = df.iloc[:, col_idx]
        
        # 存储已见内容及其编号
        seen_content = {}
        
        for row_idx, content in enumerate(col_data):
            if pd.isna(content) or str(content).strip() == "":
                # 空内容保持为空
                result_df.iloc[row_idx, col_idx] = ""
                # 空内容不更新父级编号
                continue
            
            # 获取当前行的父级编号
            parent_code = parent_codes[row_idx]
            
            # 生成唯一键用于识别相同父级下的相同内容
            content_key = (parent_code, str(content))
            
            if content_key not in seen_content:
                if col_idx == 0:
                    # 第一级：使用字母编号
                    counter = len(seen_content)
                    # 将数字转换为字母序列
                    code = number_to_letters(counter + 1)
                else:
                    # 其他级：数字编号
                    if parent_code not in counters[col_idx]:
                        counters[col_idx][parent_code] = 0
                    
                    counters[col_idx][parent_code] += 1
                    num = counters[col_idx][parent_code]
                    
                    if col_idx == 1:
                        # 第二级：直接数字
                        code = f"{parent_code}{num}"
                    else:
                        # 第三级及以后：使用连字符
                        code = f"{parent_code}-{num}"
                
                seen_content[content_key] = code
            else:
                code = seen_content[content_key]
            
            # 更新结果
            result_df.iloc[row_idx, col_idx] = f"{code} {content}"
            
            # 更新父级编号为当前编号
            parent_codes[row_idx] = code
    
    return result_df

def build_replacement_dict_from_sheet1(sheet1_df, target_col='对应特征'):
    """
    从处理后的Sheet1构建替换字典
    
    参数:
    sheet1_df: DataFrame, 处理后的Sheet1数据
    target_col: str, 目标列名称
    
    返回:
    dict: 原始名称 -> 新编号名称 的映射字典
    """
    replacement_dict = {}
    
    # 确定需要处理的列
    if target_col in sheet1_df.columns:
        target_idx = sheet1_df.columns.get_loc(target_col)
        num_cols = target_idx
    else:
        num_cols = len(sheet1_df.columns)
    
    # 处理每一列
    for col_idx in range(num_cols):
        col_data = sheet1_df.iloc[:, col_idx]
        
        for content in col_data:
            if pd.isna(content) or str(content).strip() == "":
                continue
            
            # 从处理后的字符串中提取原始名称
            # 格式是 "编号 原始名称"，例如 "A (Ⅰ)基础地理"
            parts = str(content).split(' ', 1)
            if len(parts) == 2:
                new_name = parts[0] + " " + parts[1]  # 处理后的名称
                original_name = parts[1]  # 原始名称
                
                # 添加到替换字典（使用原始名称作为键）
                if original_name not in replacement_dict:
                    replacement_dict[original_name] = new_name
                else:
                    # 如果已经存在，使用更具体的（通常是更长的路径）
                    if len(new_name) > len(replacement_dict[original_name]):
                        replacement_dict[original_name] = new_name
    
    return replacement_dict

def process_sheet2_with_replacement(sheet2_df, replacement_dict):
    """
    使用替换字典处理Sheet2
    
    参数:
    sheet2_df: DataFrame, Sheet2数据
    replacement_dict: dict, 替换字典
    
    返回:
    DataFrame: 处理后的Sheet2
    """
    result_df = sheet2_df.copy()
    
    # 遍历Sheet2的每个单元格
    for col in sheet2_df.columns:
        for row_idx, cell_value in enumerate(sheet2_df[col]):
            if pd.isna(cell_value):
                continue
            
            cell_str = str(cell_value)
            
            # 如果单元格包含多个用顿号分隔的部分，需要分别处理
            if "、" in cell_str:
                parts = cell_str.split("、")
                replaced_parts = []
                
                for part in parts:
                    part = part.strip()
                    # 尝试完全匹配
                    if part in replacement_dict:
                        replaced_parts.append(replacement_dict[part])
                    else:
                        # 如果没有完全匹配，尝试部分匹配
                        replaced = part
                        # 按长度排序，先尝试匹配较长的键
                        for original, new in sorted(replacement_dict.items(), key=lambda x: len(x[0]), reverse=True):
                            if original in part:
                                # 替换部分字符串
                                replaced = replaced.replace(original, new)
                                break
                        replaced_parts.append(replaced)
                
                # 重新组合
                result_df.at[row_idx, col] = "、".join(replaced_parts)
            else:
                # 单个部分
                if cell_str in replacement_dict:
                    result_df.at[row_idx, col] = replacement_dict[cell_str]
                else:
                    # 尝试部分匹配
                    replaced = cell_str
                    # 按长度排序，先尝试匹配较长的键
                    for original, new in sorted(replacement_dict.items(), key=lambda x: len(x[0]), reverse=True):
                        if original in cell_str:
                            replaced = replaced.replace(original, new)
                            break
                    result_df.at[row_idx, col] = replaced
    
    return result_df

def process_excel_with_merged_cells(input_file, sheet1_name='Sheet1', sheet2_name='Sheet2', 
                                    target_col='对应特征', output_file=None):
    """
    完整的Excel处理流程：
    1. 拆分合并单元格
    2. 对Sheet1进行编号
    3. 用Sheet1的编号替换Sheet2中的内容
    
    参数:
    input_file: str, 输入Excel文件路径
    sheet1_name: str, Sheet1的工作表名称
    sheet2_name: str, Sheet2的工作表名称
    target_col: str, 目标列名称
    output_file: str, 输出Excel文件路径
    
    返回:
    tuple: (处理后的Sheet1, 处理后的Sheet2, 输出文件路径)
    """
    print("="*80)
    print("开始处理Excel文件")
    print(f"输入文件: {input_file}")
    print(f"Sheet1名称: {sheet1_name}")
    print(f"Sheet2名称: {sheet2_name}")
    print(f"目标列: {target_col}")
    print("="*80)
    
    # 步骤1: 处理合并单元格
    print("\n步骤1: 处理合并单元格")
    print("-"*40)
    
    # 创建临时文件处理合并单元格
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
        temp_path = tmp_file.name
    
    try:
        # 填充合并单元格
        fill_merged_cells(input_file, temp_path, sheet_names=[sheet1_name, sheet2_name])
        
        # 步骤2: 读取处理后的数据
        print("\n步骤2: 读取处理后的数据")
        print("-"*40)
        
        try:
            sheet1_original = pd.read_excel(temp_path, sheet_name=sheet1_name)
            sheet2_original = pd.read_excel(temp_path, sheet_name=sheet2_name)
        except Exception as e:
            # 如果指定的工作表名称不存在，尝试使用默认名称
            print(f"警告: 无法读取指定的工作表: {e}")
            print("尝试读取所有可用工作表...")
            all_sheets = pd.read_excel(temp_path, sheet_name=None)
            available_sheets = list(all_sheets.keys())
            print(f"可用的工作表: {available_sheets}")
            
            if sheet1_name in all_sheets:
                sheet1_original = all_sheets[sheet1_name]
            elif available_sheets:
                sheet1_original = all_sheets[available_sheets[0]]
                sheet1_name = available_sheets[0]
                print(f"使用 '{sheet1_name}' 作为Sheet1")
            
            if sheet2_name in all_sheets:
                sheet2_original = all_sheets[sheet2_name]
            elif len(available_sheets) > 1:
                sheet2_original = all_sheets[available_sheets[1]]
                sheet2_name = available_sheets[1]
                print(f"使用 '{sheet2_name}' 作为Sheet2")
            else:
                raise ValueError(f"文件中只有一个工作表，但需要至少两个工作表")
        
        print(f"Sheet1 '{sheet1_name}' 大小: {sheet1_original.shape}")
        print(f"Sheet2 '{sheet2_name}' 大小: {sheet2_original.shape}")
        
        # 步骤3: 处理Sheet1
        print("\n步骤3: 处理Sheet1 (生成编号)")
        print("-"*40)
        
        sheet1_processed = process_hierarchical_table_generic(sheet1_original, target_col)
        
        print("Sheet1处理完成!")
        print("前5行示例:")
        print(sheet1_processed.head())
        
        # 步骤4: 构建替换字典
        print("\n步骤4: 构建替换字典")
        print("-"*40)
        
        replacement_dict = build_replacement_dict_from_sheet1(sheet1_processed, target_col)
        
        print(f"构建了 {len(replacement_dict)} 个替换项")
        print("替换字典前10个示例:")
        for i, (original, new) in enumerate(list(replacement_dict.items())[:10]):
            print(f"  {original} -> {new}")
        if len(replacement_dict) > 10:
            print(f"  ... 还有 {len(replacement_dict) - 10} 项")
        
        # 步骤5: 处理Sheet2
        print("\n步骤5: 处理Sheet2 (应用替换)")
        print("-"*40)
        
        sheet2_processed = process_sheet2_with_replacement(sheet2_original, replacement_dict)
        
        print("Sheet2处理完成!")
        print("Sheet2前5行示例:")
        print(sheet2_processed.head())
        
        # 步骤6: 保存结果
        print("\n步骤6: 保存结果")
        print("-"*40)
        
        # 生成输出文件名
        if output_file is None:
            if input_file.endswith('.xlsx'):
                output_file = input_file.replace('.xlsx', '_processed.xlsx')
            elif input_file.endswith('.xls'):
                output_file = input_file.replace('.xls', '_processed.xlsx')
            else:
                output_file = input_file + '_processed.xlsx'
        
        # 保存到Excel文件
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            sheet1_processed.to_excel(writer, sheet_name=f'{sheet1_name}', index=False)
            # sheet2_original.to_excel(writer, sheet_name=f'{sheet2_name}_原始', index=False)
            sheet2_processed.to_excel(writer, sheet_name=f'{sheet2_name}', index=False)
        
        print(f"\n处理完成! 结果已保存到: {output_file}")
        print("包含的工作表:")
        print(f"  1. {sheet1_name}_编号后: 处理后的Sheet1")
        print(f"  2. {sheet2_name}_原始: 原始的Sheet2")
        print(f"  3. {sheet2_name}_替换后: 替换后的Sheet2")
        
        return sheet1_processed, sheet2_processed, output_file
    
    finally:
        # 删除临时文件
        if os.path.exists(temp_path):
            os.unlink(temp_path)
            print(f"\n已删除临时文件: {temp_path}")

# def main():
#     """
#     主函数，演示完整处理流程
#     """
#     # 创建示例数据
#     print("="*80)
#     print("创建示例数据...")
#     print("="*80)
    
#     # Sheet1数据 (模拟有合并单元格的情况)
#     sheet1_data = {
#         '一级分类': [
#             '(Ⅰ)基础地理', '(Ⅰ)基础地理', '(Ⅰ)基础地理', '(Ⅰ)基础地理', 
#             '(Ⅰ)基础地理', '(Ⅰ)基础地理', '(Ⅱ)调查规划', '(Ⅱ)调查规划', 
#             '(Ⅱ)调查规划', '(Ⅱ)调查规划'
#         ],
#         '二级分类': [
#             '(A)定位基础', '(A)定位基础', '(B)遥感资料', '(B)遥感资料', 
#             '(B)遥感资料', '(C)基础测绘产品', '(F)调查监测', '(F)调查监测', 
#             '(F)调查监测', '(F)调查监测'
#         ],
#         '三级分类': [
#             '(1)测量控制点（CPTP）', '(1)测量控制点（CPTP）', '(2)卫星影像', 
#             '(3)航空影像', '(4)点云数据', '(5)数字高程模型（DEM)', 
#             '(18)土地资源', '(18)土地资源', '(19)地下资源', '(19)地下资源'
#         ],
#         '四级分类': [
#             '(1.1)平面控制点', '(1.2)高程控制点', '(2.1)光学影像', 
#             '', '', '(5.1)1米格网DEM（基础版）', 
#             '(18.5)城乡建设用地现状调查', '(18.5)城乡建设用地现状调查', 
#             '(19.1)地质调查', '(19.1)地质调查'
#         ],
#         '对应特征': [
#             '', '', '全色、亚米、多光谱、高光谱', '', '', '', 
#             '公共服务设施', '交通设施', '地质矿产调查评价', '城市地质调查'
#         ]
#     }
    
#     # Sheet2数据 (包含多个分类的文本)
#     sheet2_data = {
#         '分类路径': [
#             "(III)业务管理、(H)资源资产、(28)不动产登记、(28.1)权籍调查成果、"
#             "(III)业务管理、(H)资源资产、(28)不动产登记、(28.2)地籍区、"
#             "(Ⅲ)业务管理、(H)资源资产、(28)不动产登记、(28.3)地籍子区、"
#             "(Ⅲ)业务管理、(H)资源资产、(28)不动产登记、(28.4)不动产单元、"
#             "(Ⅲ)业务管理、(H)资源资产、(28)不动产登记、(28.5)业务管理、"
#             "(H)资源资产、(28)不动产权利、(Ⅲ)业务管理、(H)资源资产、"
#             "(28)不动产权利人、(Ⅲ)业务管理、(H)资源资产、(29)自然资源确权登记、"
#             "(29.1)自然资源登记单元、(Ⅲ)业务管理、(H)资源资产、"
#             "(29)自然资源确权登记、(29.2)全民所有自然资源权利主体、"
#             "(Ⅲ)业务管理、(H)资源资产、(30)争议调处、(30.1)土地、林地权属争议调处、"
#             "(Ⅲ)业务管理、(J)管制利用、(36)海洋管理、(36.1)海域海岛管理"
#         ],
#         '其他信息': ['示例单元格内容']
#     }
    
#     # 创建示例Excel文件
#     example_input = "示例文件_原始.xlsx"
    
#     with pd.ExcelWriter(example_input, engine='openpyxl') as writer:
#         # 写入Sheet1
#         sheet1_df = pd.DataFrame(sheet1_data)
#         sheet1_df.to_excel(writer, sheet_name='Sheet1', index=False)
        
#         # 写入Sheet2
#         sheet2_df = pd.DataFrame(sheet2_data)
#         sheet2_df.to_excel(writer, sheet_name='Sheet2', index=False)
    
#     print(f"已创建示例Excel文件: {example_input}")
    
#     # 使用完整的处理流程
#     print("\n" + "="*80)
#     print("开始完整处理流程")
#     print("="*80)
    
#     try:
#         sheet1_processed, sheet2_processed, output_file = process_excel_with_merged_cells(
#             input_file=example_input,
#             sheet1_name='Sheet1',
#             sheet2_name='Sheet2',
#             target_col='对应特征',
#             output_file='示例文件_处理结果.xlsx'
#         )
        
#         print("\n" + "="*80)
#         print("处理结果摘要")
#         print("="*80)
        
#         # 显示Sheet1的处理结果
#         print("\nSheet1处理结果 (前5行):")
#         print(sheet1_processed.head())
        
#         # 显示Sheet2的处理结果
#         print("\nSheet2处理结果:")
#         print(sheet2_processed)
        
#         # 显示一些替换示例
#         print("\n替换示例:")
#         original_text = sheet2_data['分类路径'][0]
#         processed_text = sheet2_processed.iloc[0, 0]
        
#         # 提取原始文本中的几个示例
#         original_parts = original_text.split("、")[:3]
#         processed_parts = processed_text.split("、")[:3]
        
#         for i, (orig, proc) in enumerate(zip(original_parts, processed_parts)):
#             if orig != proc:
#                 print(f"  示例{i+1}:")
#                 print(f"    原始: {orig}")
#                 print(f"    替换后: {proc}")
        
#         print("\n" + "="*80)
#         print("处理完成!")
#         print("="*80)
        
#     except Exception as e:
#         print(f"处理过程中发生错误: {e}")
#         import traceback
#         traceback.print_exc()

# if __name__ == "__main__":
#     main()



sheet1_processed, sheet2_processed, output_file = process_excel_with_merged_cells(
    # "interpretation_specification/src/other/广州市规划和自然资源政务数据资源目录.xlsx",
    "/usr/local/app/volume/kb_cls_grad/interpretation_specification/src/other/广州市规划和自然资源政务数据资源目录.xlsx",
    sheet1_name='Sheet1',  # Sheet1的工作表名称
    sheet2_name='Sheet2',  # Sheet2的工作表名称
    target_col='对应特征',   # 目标列名称
    output_file='interpretation_specification/src/other/处理结果.xlsx'  # 可选，输出文件路径
)
