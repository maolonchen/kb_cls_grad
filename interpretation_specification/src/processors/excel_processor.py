# -*- coding: utf-8 -*-
"""
excel转json

本模块用于将Excel文件转换为具有完整树形结构的JSON数据。
主要功能包括：
1. 去除Excel中的合并单元格，并填充内容
2. 将处理后的Excel文件解析为树状JSON结构
"""


# ####################################### xls去除合并单元格 ######################################
import openpyxl
from openpyxl.utils import range_boundaries
import os
import re
import json
from rich import print
import pandas as pd
from openpyxl import load_workbook
from config.settings import XLS_ORI_PATH, XLS_UNCOMBINE_PATH, JSON_ORI_PATH
from config.settings import get_dynamic_xls_path

# xls去除合并单元格
def fill_merged_cells(file_path, output_path, sheet_names=None):
    """
    填充Excel工作簿中指定工作表的合并单元格
    
    参数:
        file_path: Excel文件路径
        sheet_names: 要处理的工作表名称列表，如果为None则处理所有工作表
    """
    # 加载工作簿
    wb = openpyxl.load_workbook(file_path)
    
    # 如果未指定工作表，则处理所有工作表
    if sheet_names is None:
        sheet_names = wb.sheetnames
    
    # 处理每个工作表
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        print(f"正在处理工作表: {sheet_name}")
        
        # 获取所有合并单元格范围（创建副本）
        merged_ranges = list(ws.merged_cells.ranges)
        
        # 遍历所有合并单元格
        for merged_range in merged_ranges:
            # 获取合并单元格的边界
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            
            # 获取主单元格的值(左上角单元格)
            main_cell = ws.cell(row=min_row, column=min_col)
            value = main_cell.value
            
            # 先取消合并
            ws.unmerge_cells(str(merged_range))
            
            # 填充合并区域中的所有单元格
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = value
    
    # 保存修改后的文件
    wb.save(output_path)
    print(f"处理完成，结果已保存为: {output_path}")
    return output_path


# ####################################### xls转json（完整树形结构） ######################################
def split_string_to_in_and_name(input_string):
    """
    拆分文件名中的索引和名称，若不符合格式返回默认值
    
    参数:
        input_string: 输入字符串，期望格式为 "数字 - 名称"
    
    返回:
        dict: 包含 'index' 和 'name' 键的字典
    """
    match = re.search(r"(\d+)\s*[-－]\s*(.+)", input_string)
    if match:
        return {'index': match.group(1), 'name': match.group(2).strip()}
    else:
        # 如果不匹配，就将整个字符串作为 name，index 留空
        return {'index': '', 'name': input_string.strip()}

def get_merged_cells_info(sheet, hidden_cols, hidden_rows):
    """
    获取工作表中的合并单元格信息
    
    参数:
        sheet: openpyxl工作表对象
        hidden_cols: 隐藏列集合
        hidden_rows: 隐藏行集合
    
    返回:
        list: 合并单元格信息列表
    """
    merged_cells_info = []
    for merged_range in sheet.merged_cells.ranges:
        start_row = merged_range.min_row - 1
        start_col = merged_range.min_col - 1
        end_row = merged_range.max_row - 1
        end_col = merged_range.max_col - 1

        # 计算大于当前start_row、start_col的隐藏行/列的个数
        hidden_rows_count_start = sum(1 for row in hidden_rows if row < start_row)
        hidden_cols_count_start = sum(1 for col in hidden_cols if col < start_col)

        # 更新去掉隐藏行列后的start_row、start_col
        fixed_start_row = start_row - hidden_rows_count_start
        fixed_start_col = start_col - hidden_cols_count_start

        # 计算去掉隐藏行列后的合并单元格的长宽
        len_row = end_row - start_row + 1 - sum(1 for row in hidden_rows if row <= end_row and row >= start_row)
        len_col = end_col - start_col + 1 - sum(1 for col in hidden_cols if col <= end_col and col >= start_col)

        if (len_row != 0 and len_col != 0):
            fixed_end_row = fixed_start_row + len_row - 1
            fixed_end_col = fixed_start_col + len_col - 1
            merged_cells_info.append({
                "start_row": fixed_start_row if fixed_start_row != -1 else 0,
                "start_col": fixed_start_col if fixed_start_col != -1 else 0,
                "end_row": fixed_end_row,
                "end_col": fixed_end_col,
                "value": sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            })

    return merged_cells_info

def process_excel_to_tree(file_path):
    """
    将Excel文件转换为完整的树形结构
    
    参数:
        file_path: Excel文件路径
    
    返回: 完整的树结构对象
    """
    # 获取文件所在目录
    root_dir = os.path.dirname(file_path)
    
    # 获取文件所在目录的名称（最后一级目录）
    root_dir_name = os.path.basename(root_dir)
    root_names = split_string_to_in_and_name(root_dir_name)
    
    # 创建根目录对象
    tree = {
        'type': 'dir',
        'index': root_names['index'],
        'name': root_names['name'],
        'path': '/root',
        'children': []  # 将包含所有表格
    }
    
    # 加载Excel文件
    workbook = load_workbook(file_path, data_only=True)
    
    # 获取文件基本信息
    base_name = os.path.basename(file_path)  # 获取文件名（带扩展名）
    base_name = os.path.splitext(base_name)[0]  # 无扩展名
    table_names = split_string_to_in_and_name(base_name)
    
    # 创建表格对象
    table = {
        'type': 'table',
        'index': table_names['index'],
        'name': table_names['name'],
        'path': f'/root/{table_names["name"]}',
        'children': []  # 将包含所有工作表
    }
    
    # 获取所有可见工作表名称
    sheet_names = [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == 'visible']
    
    # 处理每个工作表
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        # 获取隐藏行和隐藏列的信息
        hidden_rows = {row - 1 for row in range(1, sheet.max_row + 1) if sheet.row_dimensions[row].hidden}
        hidden_cols = {ord(col) - ord('A') for col in sheet.column_dimensions if sheet.column_dimensions[col].hidden}

        # 使用Pandas读取Sheet数据
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        # 去掉隐藏的行和列
        df = df.drop(list(hidden_rows), axis=0)  # 去掉隐藏的行
        df = df.drop(list(hidden_cols), axis=1)  # 去掉隐藏的列
        # 将列号重新排序
        df.columns = range(len(df.columns))

        # 将每一行数据转为字典
        sheet_data = df.to_dict(orient="records")

        # 清理字典中的NaN值
        cleaned_sheet_data = []
        for record in sheet_data:
            cleaned_record = {str(key): value for key, value in record.items() if not pd.isna(value)}
            cleaned_sheet_data.append(cleaned_record)

        # 获取合并单元格信息
        merged_cells_info = get_merged_cells_info(sheet, hidden_cols, hidden_rows)

        # 创建工作表对象
        sheet_obj = {
            'type': 'sheet',
            'name': sheet_name,
            'path': f'/root/{table_names["name"]}/{sheet_name}',
            'info': {
                "data": cleaned_sheet_data,
                "merged_cells": merged_cells_info
            }
        }
        
        # 将工作表添加到表格
        table['children'].append(sheet_obj)
    
    # 将表格添加到目录
    tree['children'].append(table)
    
    return tree


def main(specification_u_id=None):
    XLS_ORI_PATH = get_dynamic_xls_path(specification_u_id)
    
    try:
        # 预处理XLS文件（去除合并单元格）
        print("开始预处理XLS文件（去除合并单元格）...")
        fill_merged_cells(XLS_ORI_PATH, output_path=XLS_UNCOMBINE_PATH, sheet_names=['Sheet1', 'Sheet2'])
        
        # 转换XLS为完整的树形结构JSON
        print("\n开始转换XLS为完整的树形结构JSON...")
        tree_structure = process_excel_to_tree(XLS_UNCOMBINE_PATH)
        
        # 保存JSON结果
        print("\n保存完整的树形结构JSON...")
        output_dir = os.path.dirname(XLS_ORI_PATH)
        output_file = JSON_ORI_PATH

        # 将树形结构写入文件
        with open(output_file, "w", encoding="utf-8") as file:
            json.dump(tree_structure, file, ensure_ascii=False, indent=4)

        print(f"完整的树形结构JSON已成功保存到: {output_file}")
        print("树形结构示例:")
        print(f"根目录: {tree_structure['name']}")
        for table in tree_structure['children']:
            print(f"  表格: {table['name']}")
            for sheet in table['children']:
                print(f"    工作表: {sheet['name']} ({len(sheet['info']['data'])}行数据)")
        
        return True
    except Exception as e:
        print(f"执行过程中出现错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--specification_u_id', type=str, help='行业标识符')
    args = parser.parse_args()
    main(args.specification_u_id)