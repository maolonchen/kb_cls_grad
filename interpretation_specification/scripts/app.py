# -*- coding: utf-8 -*-
"""
规范解读任务接口
"""
import os
import datetime
from typing import List, Optional
from fastapi import FastAPI, UploadFile, File, Form, status
from fastapi.responses import JSONResponse, FileResponse
from fastapi import APIRouter
import pandas as pd
from openpyxl import load_workbook
import urllib.parse
import sys
import subprocess
import threading
import time
import asyncio
import json
from kafka import KafkaProducer

project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)

from interpretation_specification.services.data_element_batch_match_service import data_element_batch_match_service
from interpretation_specification.schemas.spec_schema import DataElementBatchMatchResponse, DataElementBatchMatchRequest
from src.processors.pre_excel import process_excel_with_merged_cells
from config.settings import PROJECT_ROOT, ENABLE_KAFKA
from src.utils.logger import setup_logger

logger = setup_logger(__name__)

# 初始化 Kafka Producer
# producer = KafkaProducer(bootstrap_servers=['kafka1:9092'])  # 内部 kafka 暂停使用
# producer = None
try:
    if ENABLE_KAFKA:
        producer = KafkaProducer(
            bootstrap_servers=['192.168.10.15:39092'],
            sasl_plain_username='sw',
            sasl_plain_password='siweicn123',
            security_protocol='SASL_PLAINTEXT',
            sasl_mechanism='PLAIN',
            value_serializer=lambda v: json.dumps(v, ensure_ascii=False).encode('utf-8')
        )
        logger.info("Kafka Producer 初始化成功")
    else:
        logger.info("Kafka 已禁用，跳过初始化（分类分级暂不使用 kafka）")
except Exception as e:
    logger.error(f"Kafka Producer 初始化失败: {e}")
    producer = None  # 确保即使初始化失败也设置为 None  ##############################################################################################

# 创建路由器实例
router = APIRouter()

def send_kafka_task_status(task_uid: str, state: int, error_msg: Optional[str] = None, push_date: str = None):
    """
    向 Kafka 发送任务状态消息
    :param task_uid: 任务唯一标识
    :param state: 任务状态 (1: 待执行, 2: 执行中, 4: 任务完成, 6: 任务失败)
    :param error_msg: 错误信息 (仅在失败时使用)
    :param push_date: 时间戳
    """
    # 如果未启用Kafka，则直接返回
    if not ENABLE_KAFKA or not producer:
        return
    
    message = {
        "PushContentType": 0,
        "PushDate": push_date or datetime.datetime.now().isoformat(),
        "TaskUId": task_uid,
        "Value": {
            "State": state,
            "ErrorMsg": error_msg
        }
    }
    
    try:
        producer.send('scip_specification_analysis_task', message)
        producer.flush()  # 添加flush确保消息发送
        logger.info(f"已发送 Kafka 消息: {message}")
    except Exception as e:
        logger.error(f"发送 Kafka 消息失败: {e}", exc_info=True)

def create_directory_structure(base_dir: str, structure: dict):
    """
    递归创建目录结构
    :param base_dir: 基础目录
    :param structure: 目录结构字典
    """
    for name, children in structure.items():
        path = os.path.join(base_dir, name)
        os.makedirs(path, exist_ok=True)
        logger.info(f"Created directory: {path}")
        if children:
            create_directory_structure(path, children)

def get_safe_filename(filename: str) -> str:
    """
    获取安全的文件名，处理中文乱码问题
    使用原始文件名但确保路径安全
    :param filename: 原始文件名
    :return: 安全的文件名
    """
    # 获取文件名（不含路径）
    basename = os.path.basename(filename)
    
    # 在Linux系统上直接使用原始文件名（支持UTF-8）
    if sys.platform != "win32":
        return basename
    
    # 在Windows系统上处理中文编码问题
    try:
        # 尝试使用GBK编码（Windows默认编码）
        return basename.encode('gbk').decode('gbk')
    except:
        # 如果GBK编码失败，使用URL编码
        return urllib.parse.quote(basename)

def get_full_path(base_dir: str, specificationUId: str, version_dir: str, file_type: str, filename: str) -> str:
    """
    构建完整的文件路径
    :param base_dir: 基础目录
    :param specificationUId: 规范唯一标识
    :param version_dir: 版本目录
    :param file_type: 文件类型
    :param filename: 文件名
    :return: 完整的文件路径
    """
    return os.path.join(
        base_dir,
        f"standard_{specificationUId}",
        "01_raw_documents",
        version_dir,
        file_type,
        filename
    )

async def save_file(file: UploadFile, target_path: str):
    """
    安全保存文件到指定路径
    :param file: 上传的文件
    :param target_path: 目标路径
    :return: 目标路径
    """
    target_dir = os.path.dirname(target_path)
    os.makedirs(target_dir, exist_ok=True)
    logger.info(f"确保目录存在: {target_dir}")
    
    # 在Windows上确保使用二进制模式写入
    try:
        with open(target_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        logger.info(f"保存文件: {target_path}")
        logger.info(f"文件大小: {len(content)} 字节")
        return target_path
    except Exception as e:
        logger.error(f"保存文件失败: {e}")
        raise

def send_grad_result_to_kafka(task_u_id: str):
    """
    读取并发送分级结果到Kafka
    :param task_u_id: 任务唯一标识
    """
    # 如果未启用Kafka，则直接返回
    if not ENABLE_KAFKA or not producer:
        return
    
    try:
        grad_file_path = os.path.join(PROJECT_ROOT, "data", "processed", "kafka_output_grad.json")
        if os.path.exists(grad_file_path):
            with open(grad_file_path, 'r', encoding='utf-8') as f:
                grad_data = json.load(f)
            
            # 为每个分级数据添加applicableSceneType字段
            add_applicable_scene_type_to_grad_data(grad_data)
            
            # 构建分级结果消息
            grad_message = {
                "PushContentType": 1,
                "PushDate": datetime.datetime.now().isoformat(),
                "TaskUId": task_u_id,
                "Value": grad_data
            }
            # 发送分级结果到Kafka
            producer.send('scip_specification_analysis_task', grad_message)
            producer.flush()  # 添加flush确保消息发送
            logger.info(f"已发送分级结果到Kafka")
        else:
            logger.warning(f"分级结果文件不存在: {grad_file_path}")
    except Exception as e:
        logger.error(f"读取或发送分级结果时出错: {str(e)}", exc_info=True)

def send_feat_result_to_kafka(task_u_id: str):
    """
    读取并发送特征结果到Kafka
    :param task_u_id: 任务唯一标识
    """
    # 如果未启用Kafka，则直接返回
    if not ENABLE_KAFKA or not producer:
        return
    
    try:
        feat_file_path = os.path.join(PROJECT_ROOT, "data", "processed", "kafka_output_feat.json")
        
        if os.path.exists(feat_file_path):
            with open(feat_file_path, 'r', encoding='utf-8') as f:
                feat_data = json.load(f)
            
            # 为每个特征项添加applicableSceneType字段
            add_applicable_scene_type_to_feat_data(feat_data)
            
            # 将数据分批，每批10个元素
            batch_size = 10
            for i in range(0, len(feat_data), batch_size):
                batch = feat_data[i:i + batch_size]
                
                # 构建特征结果消息
                feat_message = {
                    "PushContentType": 3,
                    "PushDate": datetime.datetime.now().isoformat(),
                    "TaskUId": task_u_id,
                    "Value": batch
                }
                
                # 发送特征结果批次到Kafka
                producer.send('scip_specification_analysis_task', feat_message)
                producer.flush()  # 添加flush确保消息发送
                logger.info(f"已发送特征结果批次 {i//batch_size + 1} 到Kafka，包含 {len(batch)} 个项目")
        else:
            logger.warning(f"特征结果文件不存在: {feat_file_path}")
    except Exception as e:
        logger.error(f"读取或发送特征结果时出错: {str(e)}", exc_info=True)
        
def add_applicable_scene_type_to_grad_data(grad_data):
    """
    为分级数据添加applicableSceneType字段
    :param grad_data: 分级数据
    """
    # 为每个分级数据添加applicableSceneType字段
    for item in grad_data:
        if "DataGrading" in item:
            data_grading = item["DataGrading"]
            if data_grading == "核心数据":
                # 核心数据，applicableSceneType为3
                item["applicableSceneType"] = 3
            elif data_grading == "重要数据":
                # 重要数据，applicableSceneType为2
                item["applicableSceneType"] = 2
            else:
                # 其他所有格式的等级（如L1, L2, 第?级等），统一设为1
                item["applicableSceneType"] = 1

def add_applicable_scene_type_to_data(item):
    """
    为数据分类项添加applicableSceneType字段
    :param item: 数据分类项
    """
    if isinstance(item, dict):
        # 处理当前节点
        if "DataGrading" in item and item["DataGrading"]:
            data_grading = item["DataGrading"]
            if data_grading == "核心数据":
                # 核心数据，applicableSceneType为3
                item["applicableSceneType"] = 3
            elif data_grading == "重要数据":
                # 重要数据，applicableSceneType为2
                item["applicableSceneType"] = 2
            else:
                # 其他所有格式的等级（如L1, L2, 第?级等），统一设为1
                item["applicableSceneType"] = 1
        
        # 递归处理子节点
        if "SubDataClassifications" in item and item["SubDataClassifications"]:
            for sub_item in item["SubDataClassifications"]:
                add_applicable_scene_type_to_data(sub_item)

def add_applicable_scene_type_to_feat_data(feat_data):
    """
    为特征数据添加applicableSceneType字段
    :param feat_data: 特征数据
    """
    final_file_path = os.path.join(PROJECT_ROOT, "data", "processed", "final.jsonl")
    
    # 读取final.jsonl文件构建数据元素到等级的映射
    data_element_to_grading = {}
    if os.path.exists(final_file_path):
        with open(final_file_path, 'r', encoding='utf-8') as f:
            for line in f:
                try:
                    entry = json.loads(line)
                    header = entry.get('header', {})
                    data = entry.get('data', {})
                    
                    # 找到"真实数据"和"等级"字段的键
                    real_data_key = None
                    grading_key = None
                    for k, v in header.items():
                        if v == "真实数据":
                            real_data_key = k
                        elif v == "等级":
                            grading_key = k
                    
                    # 如果找到了这两个字段，建立映射关系
                    if real_data_key and grading_key:
                        real_data = data.get(real_data_key, [])
                        grading = data.get(grading_key, "")
                        
                        # 如果real_data是列表，为每个元素建立映射
                        if isinstance(real_data, list):
                            for element in real_data:
                                if element not in data_element_to_grading:
                                    data_element_to_grading[element] = set()
                                data_element_to_grading[element].add(grading)
                        # 如果real_data是字符串，建立单个映射
                        elif isinstance(real_data, str):
                            if real_data not in data_element_to_grading:
                                data_element_to_grading[real_data] = set()
                            data_element_to_grading[real_data].add(grading)
                except json.JSONDecodeError:
                    continue
    
    # 为每个特征项添加applicableSceneType字段
    for item in feat_data:
        if "DataElement" in item:
            data_element = item["DataElement"]
            applicable_scene_types = set()
            
            # 查找该数据元素对应的等级
            if data_element in data_element_to_grading:
                gradings = data_element_to_grading[data_element]
                for grading in gradings:
                    if grading == "核心数据":
                        # 核心数据，applicableSceneType包含3
                        applicable_scene_types.add(3)
                    elif grading == "重要数据":
                        # 重要数据，applicableSceneType包含2
                        applicable_scene_types.add(2)
                    else:
                        # 其他所有格式的等级（如L1, L2, 第?级等），统一添加1
                        applicable_scene_types.add(1)
            
            # 将set转换为排序后的list
            item["applicableSceneType"] = sorted(list(applicable_scene_types))
                
def send_data_result_to_kafka(task_u_id: str):
    """
    读取并发送数据分类结果到Kafka
    :param task_u_id: 任务唯一标识
    """
    # 如果未启用Kafka，则直接返回
    if not ENABLE_KAFKA or not producer:
        return
    
    try:
        data_file_path = os.path.join(PROJECT_ROOT, "data", "processed", "kafka_output_data.json")
        if os.path.exists(data_file_path):
            with open(data_file_path, 'r', encoding='utf-8') as f:
                data_data = json.load(f)
            
            # 为每个数据分类项添加applicableSceneType字段
            for item in data_data:
                add_applicable_scene_type_to_data(item)
            
            # 为每个顶级元素创建单独的消息
            for i, data_item in enumerate(data_data):
                # 构建数据分类结果消息
                data_message = {
                    "PushContentType": 2,
                    "PushDate": datetime.datetime.now().isoformat(),
                    "TaskUId": task_u_id,
                    "Value": [data_item]  # 包装成数组以匹配模板格式
                }
                
                # 发送数据分类结果到Kafka
                producer.send('scip_specification_analysis_task', data_message)
                producer.flush()  # 添加flush确保消息发送
                logger.info(f"已发送数据分类结果项 {i+1} 到Kafka")
        else:
            logger.warning(f"数据分类结果文件不存在: {data_file_path}")
    except Exception as e:
        logger.error(f"读取或发送数据分类结果时出错: {str(e)}", exc_info=True)
        
def post_process_output_files():
    """
    后处理生成的输出文件，添加applicableSceneType参数
    """
    try:
        # 处理分级结果文件
        grad_file_path = os.path.join(PROJECT_ROOT, "data", "processed", "kafka_output_grad.json")
        if os.path.exists(grad_file_path):
            with open(grad_file_path, 'r', encoding='utf-8') as f:
                grad_data = json.load(f)
            
            # 为每个分级数据添加applicableSceneType字段
            add_applicable_scene_type_to_grad_data(grad_data)
            
            # 保存更新后的数据
            with open(grad_file_path, 'w', encoding='utf-8') as f:
                json.dump(grad_data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"已更新分级结果文件: {grad_file_path}")
        
        # 处理数据分类结果文件
        data_file_path = os.path.join(PROJECT_ROOT, "data", "processed", "kafka_output_data.json")
        if os.path.exists(data_file_path):
            with open(data_file_path, 'r', encoding='utf-8') as f:
                data_data = json.load(f)
            
            # 为每个数据分类项添加applicableSceneType字段
            for item in data_data:
                add_applicable_scene_type_to_data(item)
            
            # 保存更新后的数据
            with open(data_file_path, 'w', encoding='utf-8') as f:
                json.dump(data_data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"已更新数据分类结果文件: {data_file_path}")
        
        # 处理特征结果文件
        feat_file_path = os.path.join(PROJECT_ROOT, "data", "processed", "kafka_output_feat.json")
        if os.path.exists(feat_file_path):
            with open(feat_file_path, 'r', encoding='utf-8') as f:
                feat_data = json.load(f)
            
            # 为每个特征项添加applicableSceneType字段
            add_applicable_scene_type_to_feat_data(feat_data)
            
            # 保存更新后的数据
            with open(feat_file_path, 'w', encoding='utf-8') as f:
                json.dump(feat_data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"已更新特征结果文件: {feat_file_path}")
            
    except Exception as e:
        logger.error(f"后处理输出文件时出错: {str(e)}", exc_info=True)

def run_processing_pipeline(specification_u_id: str, task_u_id: str, excel_file_path: str = None):
    """
    在后台运行处理流程并记录输出到日志文件
    :param specification_u_id: 规范唯一标识
    :param task_u_id: 任务唯一标识
    :param excel_file_path: 预处理后的Excel文件路径
    """
    try:
        # 构建日志目录
        log_dir = os.path.join(PROJECT_ROOT, "logs")
        os.makedirs(log_dir, exist_ok=True)
        
        # 构建日志文件路径，基于规范ID和当前时间
        current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file_path = os.path.join(log_dir, f"processing_{task_u_id}_{current_time}.log")
        
        # 构建命令 - 修改为正确的模块路径
        cmd = [
            sys.executable, "-m", "src.main", 
            "--all", 
            "--specification-u-id", specification_u_id
        ]
        
        # 如果提供了预处理后的Excel文件路径，可以将其作为参数传递
        if excel_file_path:
            cmd.extend(["--excel-path", excel_file_path])
        
        logger.info(f"启动处理流程: {' '.join(cmd)}")
        logger.info(f"日志文件: {log_file_path}")
        
        # 打开日志文件以写入标准输出和错误输出
        with open(log_file_path, "w") as log_file:
            # 在子进程中运行命令，并将 stdout 和 stderr 写入日志文件
            process = subprocess.Popen(
                cmd,
                cwd=PROJECT_ROOT,
                stdout=log_file,
                stderr=subprocess.STDOUT,  # 将 stderr 合并到 stdout
                text=True
            )
            
            # 等待进程完成
            process.wait()
        
        if process.returncode == 0:
            logger.info(f"处理流程执行成功: {specification_u_id}")
            
            # 后处理：为生成的文件添加applicableSceneType参数
            post_process_output_files()
            
            # 读取生成的分级结果文件并发送到Kafka
            send_grad_result_to_kafka(task_u_id)
                
            # 读取生成的数据分类结果文件并逐个发送到Kafka
            send_data_result_to_kafka(task_u_id)
            
            # 读取生成的特征结果文件并分批发送到Kafka
            send_feat_result_to_kafka(task_u_id)
                
            # 发送任务完成状态
            send_kafka_task_status(task_u_id, state=4)
        else:
            logger.error(f"处理流程执行失败: {specification_u_id}")
            send_kafka_task_status(task_u_id, state=6, error_msg="处理流程执行失败")
            
    except Exception as e:
        logger.error(f"执行处理流程时出错: {str(e)}", exc_info=True)
        send_kafka_task_status(task_u_id, state=6, error_msg=str(e))


def validate_excel_format(file_path):
    """
    验证Excel文件格式是否符合要求
    返回 (is_valid, error_messages)
    """
    errors = []
    
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        
        # 检查工作表名称
        sheet_names = wb.sheetnames
        if 'Sheet1' not in sheet_names:
            errors.append("表一的名字不是'Sheet1'")
        if 'Sheet2' not in sheet_names:
            errors.append("表二的名字不是'Sheet2'")
        
        # 即使工作表名称有问题，也要尝试检查工作表内容，所以不提前返回
        
        # 检查Sheet1的第一行内容
        if 'Sheet1' not in wb.sheetnames:
            errors.append("缺少Sheet1工作表")
        else:
            ws1 = wb['Sheet1']
            first_row_values = [cell.value for cell in ws1[1]]
            
            if not first_row_values or len(first_row_values) < 2:
                errors.append("Sheet1第一行至少需要包含'一级分类'和'对应特征'两列")
            else:
                # 检查是否以"一级分类"开头
                if not first_row_values[0] or str(first_row_values[0]) != "一级分类":
                    errors.append(f"Sheet1第一行第一列应该是'一级分类'，实际是'{first_row_values[0] if first_row_values[0] else '空值'}'")
                
                # 检查是否按"一级分类"、"二级分类"...递增的形式排列直到最后一列是"对应特征"
                # 从第二列开始检查到倒数第二列（最后一列是"对应特征"）
                for level_idx in range(1, len(first_row_values) - 1):
                    cell_value = first_row_values[level_idx]
                    
                    # 将阿拉伯数字转换为中文数字以便比较
                    arabic_to_chinese = {1: "一", 2: "二", 3: "三", 4: "四", 5: "五", 
                                        6: "六", 7: "七", 8: "八", 9: "九", 10: "十",
                                        11: "十一", 12: "十二", 13: "十三", 14: "十四", 15: "十五",
                                        16: "十六", 17: "十七", 18: "十八", 19: "十九", 20: "二十"}
                    
                    if level_idx + 1 in arabic_to_chinese:
                        expected_chinese_level = arabic_to_chinese[level_idx + 1]
                    else:
                        # 对于超过20的情况，简单处理
                        expected_chinese_level = f"{level_idx + 1}"
                    
                    expected_level = f"{expected_chinese_level}级分类"
                    if cell_value is None:
                        errors.append(f"Sheet1第一行第{level_idx+1}列为空，期望为'{expected_level}'")
                    elif str(cell_value) != expected_level:
                        errors.append(f"Sheet1第一行第{level_idx+1}列应该是'{expected_level}'，实际是'{cell_value}'")
                
                # 检查最后一列是否是"对应特征"
                if len(first_row_values) > 0:
                    last_column = first_row_values[-1]
                    if last_column != "对应特征":
                        errors.append(f"Sheet1第一行最后一列应该是'对应特征'，实际是'{last_column}'")
        
        # 检查Sheet2的第一行内容
        if 'Sheet2' not in wb.sheetnames:
            errors.append("缺少Sheet2工作表")
        else:
            ws2 = wb['Sheet2']
            sheet2_first_row = [cell.value for cell in ws2[1]]
            
            if len(sheet2_first_row) < 2:
                errors.append("Sheet2第一行至少需要'类别'和'子类及范围'两列")
            else:
                if sheet2_first_row[0] != "类别":
                    errors.append(f"Sheet2第一行第一列应该是'类别'，实际是'{sheet2_first_row[0] if sheet2_first_row[0] else '空值'}'")
                if sheet2_first_row[1] != "子类及范围":
                    errors.append(f"Sheet2第一行第二列应该是'子类及范围'，实际是'{sheet2_first_row[1] if sheet2_first_row[1] else '空值'}'")
        
        return len(errors) == 0, errors
    
    except Exception as e:
        errors.append(f"读取Excel文件时发生错误: {str(e)}")
        return False, errors


# 处理任务创建接口
@router.post("/api/v1/specification/tasks",
          responses={
              400: {"description": "无效参数或文件错误"},
              500: {"description": "服务器内部错误"}
          })

async def create_specification_task(
    taskUId: str = Form(..., description="任务唯一标识"),
    specificationUId: str = Form(..., description="规范唯一标识"),
    specificationName: str = Form(..., description="规范名称"),
    excelFile: UploadFile = File(..., description="Excel规范文件"),
    supplementFiles: Optional[List[UploadFile]] = File(None, description="补充文档(PDF/Word)")
):
    """
    创建规范解读任务接口
    - 接收任务参数和文件
    - 返回任务接收状态
    - 实际处理通过后台消息队列进行
    """
    
    # 记录上传时间
    upload_time = datetime.datetime.now().isoformat()
    
    # 发送任务状态: 待执行
    send_kafka_task_status(taskUId, state=1, push_date=upload_time)

    # 验证任务唯一标识格式
    if not taskUId or len(taskUId) < 6:
        send_kafka_task_status(taskUId, state=6, error_msg="无效的任务唯一标识")
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "code": 400, "msg": "无效的任务唯一标识"}
        )
    
    # 验证Excel文件格式
    excel_ext = os.path.splitext(excelFile.filename)[1].lower()
    if excel_ext not in [".xls", ".xlsx"]:
        send_kafka_task_status(taskUId, state=6, error_msg="无效的Excel文件格式")
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "code": 400, "msg": "无效的Excel文件格式"}
        )
    
    # 验证补充文件格式
    if supplementFiles:
        for file in supplementFiles:
            ext = os.path.splitext(file.filename)[1].lower()
            if ext not in [".pdf", ".doc", ".docx"]:
                send_kafka_task_status(taskUId, state=6, error_msg=f"无效的补充文件格式: {file.filename}")
                return JSONResponse(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    content={"success": False, "code": 400, "msg": f"无效的补充文件格式: {file.filename}"}
                )
    
    try:
        # 基础目录结构
        base_dir = os.path.join(PROJECT_ROOT, "classification_and_grading")
        
        # 完整的目录结构定义
        directory_structure = {
            f"standard_{specificationUId}": {
                "01_raw_documents": {
                    "excel": {},  # Excel文件存放目录
                    "supplements": {}  # 补充文件存放目录
                }
            }
        }
        
        # 创建所有必要的目录
        create_directory_structure(base_dir, directory_structure)
        
        # 构建Excel文件路径 - 使用原始文件名
        excel_filename = get_safe_filename(excelFile.filename)
        excel_path = os.path.join(
            base_dir,
            f"standard_{specificationUId}",
            "01_raw_documents",
            "excel",
            excel_filename
        )
        
        # 保存Excel文件
        saved_excel_path = await save_file(excelFile, excel_path)
        
        # 在验证Excel文件格式之前，先对Excel文件进行预处理
        logger.info(f"开始对Excel文件进行预处理: {saved_excel_path}")
        
        # 使用Excel转换逻辑处理文件
        processed_excel_path = saved_excel_path  # 使用相同的路径，覆盖原始文件
        
        # 调用Excel处理函数
        try:
            sheet1_processed, sheet2_processed, processed_excel_path = process_excel_with_merged_cells(
                input_file=saved_excel_path,
                output_file=processed_excel_path  # 保存到相同路径，覆盖原始文件
            )
            logger.info(f"Excel文件预处理完成并已覆盖原始文件: {processed_excel_path}")
            
            # 确保预处理后的文件确实被保存
            if not os.path.exists(processed_excel_path):
                logger.error(f"预处理后的文件未找到: {processed_excel_path}")
                send_kafka_task_status(taskUId, state=6, error_msg="Excel文件预处理失败，文件未保存")
                return JSONResponse(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    content={"success": False, "code": 400, "msg": "Excel文件预处理失败，文件未保存"}
                )
            
            # 生成映射关系JSON文件
            mapping_json_path = os.path.join(
                project_root, 
                "src", 
                "other", 
                f"mapping_{specificationUId}.json"
            )
            
            # 从处理后的数据中提取映射关系
            mapping_dict = {}
            
            # 从Sheet1中提取映射关系 (原始内容 -> 编号后内容)
            if hasattr(sheet1_processed, 'columns'):
                for col_idx in range(len(sheet1_processed.columns)):
                    col_data = sheet1_processed.iloc[:, col_idx]
                    for _, cell_value in col_data.items():
                        if pd.notna(cell_value) and cell_value != '':
                            # 检查是否是带有编号的格式，如 "A (Ⅰ)基础地理"
                            if ' ' in str(cell_value):
                                parts = str(cell_value).split(' ', 1)
                                if len(parts) == 2:
                                    numbered_part = parts[0]
                                    original_part = parts[1]
                                    # 添加原始内容到编号后内容的映射
                                    mapping_dict[original_part] = str(cell_value)
            
            # 保存映射关系到JSON文件
            with open(mapping_json_path, 'w', encoding='utf-8') as f:
                json.dump(mapping_dict, f, ensure_ascii=False, indent=2)
            
            logger.info(f"映射关系已保存到: {mapping_json_path}")
            
        except Exception as e:
            logger.error(f"Excel文件预处理失败: {str(e)}")
            send_kafka_task_status(taskUId, state=6, error_msg=f"Excel文件预处理失败: {str(e)}")
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content={"success": False, "code": 400, "msg": f"Excel文件预处理失败: {str(e)}"}
            )
        
        # 验证Excel文件格式（使用处理后的文件）
        is_valid, validation_errors = validate_excel_format(processed_excel_path)
        if not is_valid:
            error_msg = "Excel文件格式不符合规范: " + "; ".join(validation_errors)
            send_kafka_task_status(taskUId, state=6, error_msg=error_msg)
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content={"success": False, "code": 400, "msg": error_msg}
            )
        
        # 保存补充文件
        saved_supp_paths = []
        if supplementFiles:
            for file in supplementFiles:
                supp_filename = get_safe_filename(file.filename)
                supp_path = os.path.join(
                    base_dir,
                    f"standard_{specificationUId}",
                    "01_raw_documents",
                    "supplements",
                    supp_filename
                )
                saved_path = await save_file(file, supp_path)
                saved_supp_paths.append(saved_path)
        
        # 记录所有保存的文件路径
        logger.info("=" * 50)
        logger.info("文件保存位置验证".center(50))
        logger.info("=" * 50)
        logger.info(f"Excel文件: {processed_excel_path}")  # 使用处理后的文件路径
        for i, path in enumerate(saved_supp_paths):
            logger.info(f"补充文件 #{i+1}: {path}")
        logger.info("=" * 50)
        
        # 验证文件是否实际保存
        if not os.path.exists(processed_excel_path):  # 验证处理后的文件
            send_kafka_task_status(taskUId, state=6, error_msg="Excel文件保存失败")
            return JSONResponse(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                content={"success": False, "code": 500, "msg": "Excel文件保存失败"}
            )
        
        # 启动后台处理流程
        logger.info(f"启动后台处理流程，specificationUId: {specificationUId}")
        # 使用线程在后台运行处理流程，避免阻塞API响应

        # 发送任务状态: 执行中
        send_kafka_task_status(taskUId, state=2)

        thread = threading.Thread(
            target=run_processing_pipeline, 
            args=(specificationUId, taskUId),
            daemon=False
        )
        thread.start()
        
        logger.info(f"Successfully created task structure for taskUId: {taskUId}")
        
        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "success": True, 
                "code": 200, 
                "msg": "任务创建成功，处理流程已在后台启动"
            }
        )
    
    except Exception as e:
        logger.error("服务器内部错误", exc_info=True)
        send_kafka_task_status(taskUId, state=6, error_msg=str(e))
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "code": 500, "msg": f"服务器内部错误: {str(e)}"}
        )


def run_processing_pipeline_for_final_jsonl(specification_u_id: str, task_u_id: str):
    """
    在后台运行处理流程，但只到生成final.jsonl文件为止
    :param specification_u_id: 规范唯一标识
    :param task_u_id: 任务唯一标识
    """
    try:
        # 构建日志目录
        log_dir = os.path.join(PROJECT_ROOT, "logs")
        os.makedirs(log_dir, exist_ok=True)
        
        # 构建日志文件路径，基于规范ID和当前时间
        current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file_path = os.path.join(log_dir, f"processing_final_jsonl_{task_u_id}_{current_time}.log")
        
        # 构建命令 - 只执行到第8步（生成final.jsonl），不执行第9步发送到Kafka
        cmd = [
            sys.executable, "-m", "src.main", 
            "--from-step", "1",
            "--specification-u-id", specification_u_id
        ]
        
        logger.info(f"启动处理流程（到final.jsonl为止）: {' '.join(cmd)}")
        logger.info(f"日志文件: {log_file_path}")
        
        # 打开日志文件以写入标准输出和错误输出
        with open(log_file_path, "w") as log_file:
            # 在子进程中运行命令，并将 stdout 和 stderr 写入日志文件
            process = subprocess.Popen(
                cmd,
                cwd=PROJECT_ROOT,
                stdout=log_file,
                stderr=subprocess.STDOUT,  # 将 stderr 合并到 stdout
                text=True
            )
            
            # 等待进程完成
            process.wait()
        
        if process.returncode == 0:
            logger.info(f"处理流程执行成功（到final.jsonl为止）: {specification_u_id}")
            
            # 发送任务完成状态
            send_kafka_task_status(task_u_id, state=4)
        else:
            logger.error(f"处理流程执行失败: {specification_u_id}")
            send_kafka_task_status(task_u_id, state=6, error_msg="处理流程执行失败")
            
    except Exception as e:
        logger.error(f"执行处理流程时出错: {str(e)}", exc_info=True)
        send_kafka_task_status(task_u_id, state=6, error_msg=str(e))

# 返回final.jsonl文件
@router.post("/api/v1/specification/tasks/final-jsonl",
          responses={
              400: {"description": "无效参数或文件错误"},
              500: {"description": "服务器内部错误"}
          })
async def create_specification_task_final_jsonl(
    taskUId: str = Form(..., description="任务唯一标识"),
    specificationUId: str = Form(..., description="规范唯一标识"),
    specificationName: str = Form(..., description="规范名称"),
    excelFile: UploadFile = File(..., description="Excel规范文件"),
    supplementFiles: Optional[List[UploadFile]] = File(None, description="补充文档(PDF/Word)")
):
    """
    创建规范解读任务接口（返回final.jsonl）
    - 接收任务参数和文件
    - 返回任务接收状态
    - 实际处理通过后台进行，但只处理到生成final.jsonl文件
    - 最终返回final.jsonl文件内容给客户端
    """
    
    # 记录上传时间
    upload_time = datetime.datetime.now().isoformat()
    
    # 发送任务状态: 待执行
    send_kafka_task_status(taskUId, state=1, push_date=upload_time)

    # 验证任务唯一标识格式
    if not taskUId or len(taskUId) < 6:
        send_kafka_task_status(taskUId, state=6, error_msg="无效的任务唯一标识")
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "code": 400, "msg": "无效的任务唯一标识"}
        )
    
    # 验证Excel文件格式
    excel_ext = os.path.splitext(excelFile.filename)[1].lower()
    if excel_ext not in [".xls", ".xlsx"]:
        send_kafka_task_status(taskUId, state=6, error_msg="无效的Excel文件格式")
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "code": 400, "msg": "无效的Excel文件格式"}
        )
    
    # 验证补充文件格式
    if supplementFiles:
        for file in supplementFiles:
            ext = os.path.splitext(file.filename)[1].lower()
            if ext not in [".pdf", ".doc", ".docx"]:
                send_kafka_task_status(taskUId, state=6, error_msg=f"无效的补充文件格式: {file.filename}")
                return JSONResponse(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    content={"success": False, "code": 400, "msg": f"无效的补充文件格式: {file.filename}"}
                )
    
    try:
        # 基础目录结构
        base_dir = os.path.join(PROJECT_ROOT, "classification_and_grading")
        
        # 完整的目录结构定义
        directory_structure = {
            f"standard_{specificationUId}": {
                "01_raw_documents": {
                    "excel": {},  # Excel文件存放目录
                    "supplements": {}  # 补充文件存放目录
                }
            }
        }
        
        # 创建所有必要的目录
        create_directory_structure(base_dir, directory_structure)
        
        # 构建Excel文件路径 - 使用原始文件名
        excel_filename = get_safe_filename(excelFile.filename)
        excel_path = os.path.join(
            base_dir,
            f"standard_{specificationUId}",
            "01_raw_documents",
            "excel",
            excel_filename
        )
        
        # 保存Excel文件
        saved_excel_path = await save_file(excelFile, excel_path)
        
        # 保存补充文件
        saved_supp_paths = []
        if supplementFiles:
            for file in supplementFiles:
                supp_filename = get_safe_filename(file.filename)
                supp_path = os.path.join(
                    base_dir,
                    f"standard_{specificationUId}",
                    "01_raw_documents",
                    "supplements",
                    supp_filename
                )
                saved_path = await save_file(file, supp_path)
                saved_supp_paths.append(saved_path)
        
        # 记录所有保存的文件路径
        logger.info("=" * 50)
        logger.info("文件保存位置验证".center(50))
        logger.info("=" * 50)
        logger.info(f"Excel文件: {saved_excel_path}")
        for i, path in enumerate(saved_supp_paths):
            logger.info(f"补充文件 #{i+1}: {path}")
        logger.info("=" * 50)
        
        # 验证文件是否实际保存
        if not os.path.exists(saved_excel_path):
            send_kafka_task_status(taskUId, state=6, error_msg="Excel文件保存失败")
            return JSONResponse(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                content={"success": False, "code": 500, "msg": "Excel文件保存失败"}
            )
        
        # 启动后台处理流程（只到生成final.jsonl为止）
        logger.info(f"启动后台处理流程（只到final.jsonl），specificationUId: {specificationUId}")
        # 使用线程在后台运行处理流程，避免阻塞API响应

        # 发送任务状态: 执行中
        send_kafka_task_status(taskUId, state=2)

        thread = threading.Thread(
            target=run_processing_pipeline_for_final_jsonl, 
            args=(specificationUId, taskUId),
            daemon=False
        )
        thread.start()
        
        logger.info(f"Successfully created task structure for taskUId: {taskUId}")
        
        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "success": True, 
                "code": 200, 
                "msg": "任务创建成功，处理流程已在后台启动（只处理到final.jsonl）"
            }
        )
    
    except Exception as e:
        logger.error("服务器内部错误", exc_info=True)
        send_kafka_task_status(taskUId, state=6, error_msg=str(e))
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "code": 500, "msg": f"服务器内部错误: {str(e)}"}
        )


@router.get("/api/v1/specification/tasks/{specificationUId}/final-jsonl",
         responses={
             200: {"description": "返回final.jsonl文件"},
             404: {"description": "文件未找到"},
             500: {"description": "服务器内部错误"}
         })
async def get_final_jsonl(specificationUId: str):
    """
    获取指定规范ID的final.jsonl文件
    - 检查文件是否存在
    - 返回final.jsonl文件内容给客户端
    """
    try:
        # 构建final.jsonl文件路径
        final_jsonl_path = os.path.join(PROJECT_ROOT, "data", "processed", "final.jsonl")
        
        # 检查文件是否存在
        if not os.path.exists(final_jsonl_path):
            logger.error(f"final.jsonl文件不存在: {final_jsonl_path}")
            return JSONResponse(
                status_code=status.HTTP_404_NOT_FOUND,
                content={
                    "success": False, 
                    "code": 404, 
                    "msg": f"final.jsonl文件不存在，规范ID: {specificationUId}"
                }
            )
        
        logger.info(f"正在返回final.jsonl文件: {final_jsonl_path}")
        
        # 返回文件
        return FileResponse(
            path=final_jsonl_path,
            filename="final.jsonl",
            media_type="application/jsonl"
        )
    
    except Exception as e:
        logger.error(f"获取final.jsonl文件时出错: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={
                "success": False, 
                "code": 500, 
                "msg": f"服务器内部错误: {str(e)}"
            }
        )


@router.post(
    "/api/v1/dataElements/match",
    response_model=DataElementBatchMatchResponse,
    summary="匹配数据特征一致性接口"
)
async def data_element_batch_match(request: DataElementBatchMatchRequest):
    """
    匹配数据特征一致性接口

    Args:
        request: 数据特征批量匹配请求参数

    Returns:
        DataElementBatchMatchResponse: 数据特征批量匹配处理结果
    """
    try:
        # 使用数据特征批量匹配服务处理请求
        result = await data_element_batch_match_service.process_data_element_batch_match(request)
        return result
    except Exception as e:
        logger.error(f"处理数据特征匹配失败: {e}", exc_info=True)
        return DataElementBatchMatchResponse(
            success=False,
            code=500,
            msg=f"处理数据特征匹配失败: {str(e)}",
            data=[]
        )



# 使用路由器创建FastAPI应用
app = FastAPI(title="规范解读任务接口")
app.include_router(router)

if __name__ == "__main__":
    import uvicorn
    
    logger.info("=" * 50)
    logger.info("启动规范解读任务服务".center(50))
    logger.info("=" * 50)
    logger.info("API 地址: http://127.0.0.1:64001")
    logger.info("文档地址: http://127.0.0.1:64001/docs")
    logger.info("文件存储目录: /classification_and_grading/")
    logger.info("=" * 50)
    
    # 创建基础存储目录
    base_dir = os.path.join(PROJECT_ROOT, "classification_and_grading")
    os.makedirs(base_dir, exist_ok=True)
    logger.info(f"基础存储目录: {base_dir}")
    
    # 获取当前文件名
    module_name = os.path.splitext(os.path.basename(__file__))[0]
    
    # 设置系统编码为UTF-8
    if sys.version_info >= (3, 7):
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    
    uvicorn.run(
        f"{module_name}:app",
        host="0.0.0.0",
        port=64001,
        reload=False,
        log_level="info",
        access_log=True,
        log_config=None
    )