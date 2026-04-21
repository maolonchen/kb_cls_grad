#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel文件处理器
用于将Excel文件转换为Markdown格式
"""

import json
import openpyxl
import logging
from pathlib import Path
import re
from difflib import SequenceMatcher
import numpy as np
from typing import Union

# 配置日志
logger = logging.getLogger(__name__)

def expand_excel_merged_cells(file_path, sheet_name=0):
    """
    读取Excel文件并展开合并的单元格
    """
    # 使用openpyxl加载工作簿以处理合并单元格
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    if isinstance(sheet_name, int):
        sheet = wb.worksheets[sheet_name]
    else:
        sheet = wb[sheet_name]
    
    # 获取工作表维度
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # 创建网格
    grid = [['' for _ in range(max_col)] for _ in range(max_row)]
    
    # 处理合并单元格
    for merged_range in sheet.merged_cells.ranges:
        min_row, min_col, max_row, max_col = (
            merged_range.min_row, 
            merged_range.min_col, 
            merged_range.max_row, 
            merged_range.max_col
        )
        
        # 获取合并单元格的值（来自第一个单元格）
        value = sheet.cell(row=min_row, column=min_col).value
        if value is None:
            value = ''
        
        # 填充合并区域
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                grid[row-1][col-1] = str(value) if value is not None else ''
    
    # 填充非合并单元格
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            # 跳过已填充的合并单元格
            if grid[row-1][col-1] == '':
                value = sheet.cell(row=row, column=col).value
                grid[row-1][col-1] = str(value) if value is not None else ''
    
    return grid

def is_numeric(value):
    """检查值是否为数字（包括带逗号的数字）"""
    if value is None or str(value).strip() == '':
        return False
    try:
        # 移除逗号并尝试转换为浮点数
        float(str(value).replace(',', ''))
        return True
    except:
        return False

def calculate_similarity(a, b):
    """计算两个字符串的相似度（0-1）"""
    if a == b:
        return 1.0
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, str(a), str(b)).ratio()

def detect_header_row(grid):
    """
    检测表头所在的行
    返回表头行索引（如果没有表头，返回None）
    """
    if not grid or len(grid) < 2:
        return None  # 至少需要2行才能判断
    
    # 计算所有数据行之间的平均相似度
    data_similarities = []
    for i in range(1, len(grid)):
        for j in range(i+1, len(grid)):
            row_i = grid[i]
            row_j = grid[j]
            similarity = 0
            comparable = 0
            
            for col_idx in range(min(len(row_i), len(row_j))):
                cell_i = str(row_i[col_idx]).strip()
                cell_j = str(row_j[col_idx]).strip()
                
                if cell_i and cell_j:
                    similarity += calculate_similarity(cell_i, cell_j)
                    comparable += 1
            
            if comparable > 0:
                data_similarities.append(similarity / comparable)
    
    # 计算数据行之间的平均相似度
    avg_data_similarity = np.mean(data_similarities) if data_similarities else 0
    
    # 尝试检测可能的表头行
    candidate_header = None
    best_score = -1
    
    # 检查每一行是否可能是表头
    # 限制只在前三行中查找表头
    max_header_row = min(3, len(grid)-1)
    for row_idx in range(max_header_row):  # 只检查前三行作为表头候选
        # 计算当前行与后续所有行的平均相似度
        header_to_data_similarities = []
        for data_idx in range(row_idx+1, len(grid)):
            row_header = grid[row_idx]
            row_data = grid[data_idx]
            similarity = 0
            comparable = 0
            
            for col_idx in range(min(len(row_header), len(row_data))):
                cell_header = str(row_header[col_idx]).strip()
                cell_data = str(row_data[col_idx]).strip()
                
                if cell_header and cell_data:
                    similarity += calculate_similarity(cell_header, cell_data)
                    comparable += 1
            
            if comparable > 0:
                header_to_data_similarities.append(similarity / comparable)
        
        # 计算当前行与数据行的平均相似度
        avg_header_to_data = np.mean(header_to_data_similarities) if header_to_data_similarities else 0
        
        # 计算得分：数据行相似度与表头-数据行相似度的差异
        score = avg_data_similarity - avg_header_to_data
        
        # 检查当前行是否像表头（文本特征）
        text_like_count = 0
        for cell in grid[row_idx]:
            cell_str = str(cell).strip()
            if cell_str and re.search(r'[a-zA-Z\u4e00-\u9fff]', cell_str) and not is_numeric(cell_str):
                text_like_count += 1
        
        # 如果文本特征明显，增加得分
        if text_like_count >= 3:
            score += 0.5
        
        logger.debug(f"行 {row_idx} 得分: {score:.2f} (数据相似度: {avg_data_similarity:.2f}, 表头-数据相似度: {avg_header_to_data:.2f}, 文本特征: {text_like_count})")
        
        # 更新最佳候选
        if score > best_score:
            best_score = score
            candidate_header = row_idx
    
    # 如果最佳得分超过阈值，则认为是表头行
    if best_score > 0.3:
        logger.info(f"检测到表头行: 行 {candidate_header}, 得分: {best_score:.2f}")
        return candidate_header
    
    logger.info("未检测到明显的表头行")
    return None

def build_tree_from_grid(grid, header_row=None):
    """
    从网格数据构建树形结构
    header_row参数指示表头所在的行索引（从0开始）
    """
    if not grid or len(grid) < 1:
        return []
    
    rows_data = []
    
    if header_row is not None:
        # 有表头的情况：使用指定行作为列名
        header_names = [str(cell).strip() for cell in grid[header_row]]
        
        # 处理数据行（从表头下一行开始）
        for row_idx in range(header_row+1, len(grid)):
            row = grid[row_idx]
            # 跳过空行
            if not any(str(cell).strip() for cell in row):
                continue
            
            # 构建当前行的数据字典
            row_data = {}
            for i, header in enumerate(header_names):
                if i < len(row):
                    row_data[header] = str(row[i]).strip()
                else:
                    row_data[header] = ""
            
            rows_data.append(row_data)
    else:
        # 无表头的情况：生成默认列名
        if len(grid) > 0:
            num_columns = len(grid[0])
            header_names = [f"column{i}" for i in range(num_columns)]
            
            # 处理所有行作为数据
            for row in grid:
                # 跳过空行
                if not any(str(cell).strip() for cell in row):
                    continue
                
                # 构建当前行的数据字典
                row_data = {}
                for i, header in enumerate(header_names):
                    if i < len(row):
                        row_data[header] = str(row[i]).strip()
                    else:
                        row_data[header] = ""
                
                rows_data.append(row_data)
    
    return rows_data

def clean_data(data):
    """
    递归清理数据中的空格：
    - 所有键的空格去除；
    - 所有字符串值的空格去除；
    - 支持嵌套字典和列表。
    """
    if isinstance(data, dict):
        return {
            str(key).replace(" ", ""): clean_data(value)
            for key, value in data.items()
        }
    elif isinstance(data, list):
        return [clean_data(value) for value in data]
    elif isinstance(data, str):
        return data.replace(" ", "")
    else:
        return data

def excel_to_jsonl(file_path, sheet_name=0):
    """
    将Excel表格转换为JSONL格式
    """
    try:
        # 处理Excel数据
        grid = expand_excel_merged_cells(file_path, sheet_name)
        
        if not grid:
            return None
        
        logger.debug(f"网格行数: {len(grid)}")
        
        # 检测表头行
        header_row = detect_header_row(grid)
        
        # 构建树形结构
        rows_data = build_tree_from_grid(grid, header_row)
        
        if not rows_data:
            return None
            
        # 转换为JSONL格式
        jsonl_lines = []
        for row in rows_data:
            # 清理整个树结构中的空格
            cleaned_row = clean_data(row)
            # 转换为JSON行
            jsonl_lines.append(json.dumps(cleaned_row, ensure_ascii=False))
        
        # 用换行符连接所有行
        return '\n'.join(jsonl_lines)
    except Exception as e:
        logger.error(f"转换Excel到JSONL时出错: {str(e)}", exc_info=True)
        return None

def process_excel_to_markdown(excel_file_path: Union[str, Path], output_path: Union[str, Path]) -> str:
    """
    将Excel文件处理为Markdown格式文件

    参数:
        excel_file_path (Union[str, Path]): Excel文件路径
        output_path (Union[str, Path]): 输出Markdown文件路径

    返回:
        str: 生成的Markdown文件路径

    异常:
        FileNotFoundError: 如果Excel文件不存在
        RuntimeError: 如果转换失败
    """
    excel_file = Path(excel_file_path)
    
    # 检查文件是否存在
    if not excel_file.exists():
        raise FileNotFoundError(f"Excel文件不存在: {excel_file_path}")
    
    # 检查文件扩展名
    if excel_file.suffix.lower() not in ['.xls', '.xlsx']:
        raise ValueError(f"不支持的文件格式: {excel_file.suffix}，仅支持.xls和.xlsx文件")
    
    try:
        # 转换Excel到JSONL
        jsonl_content = excel_to_jsonl(str(excel_file))
        
        if jsonl_content is None:
            raise RuntimeError("Excel转换为JSONL失败")
        
        # 生成Markdown内容
        md_content = f"# {excel_file.stem}\n\n"
        md_content += "```\n"
        md_content += jsonl_content
        md_content += "\n```\n"
        
        # 写入Markdown文件
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(md_content)
        
        logger.info(f"已处理Excel文件: {excel_file} -> {output_file}")
        return str(output_file)
        
    except Exception as e:
        logger.error(f"处理Excel文件 {excel_file} 时出错: {e}", exc_info=True)
        raise RuntimeError(f"处理失败: {str(e)}")